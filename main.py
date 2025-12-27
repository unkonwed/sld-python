import pythoncom
import win32com.client
import openpyxl
import pySW
from swconst import constants
import pySldWrap.sw_tools as sw_tools  # 导入pySldWrap库


表格名称 = "数据存储"  # 所生成表格的名称，注意，重名会自动删除之前的表格
版本号 = 2022  # 所用的solidworks的版本号

# 创建一个全局的数组用于存储之后获取的装配体中的零件名称。
selectedpart = []


# 创建一个遍历装配体的函数
def traverse_node(node):

    global selectedpart
    nodeObject = node.Object
    if nodeObject is not None:
        if nodeObject.GetSuppression != constants.swComponentSuppressed:
            # print(nodeObject.Name2)
            selectedpart.append(nodeObject.Name2)
            # print(selectedpart)
    childNode = node.GetFirstChild
    while childNode is not None:
        if childNode.ObjectType == constants.swFeatureManagerItem_Component:
            traverse_node(childNode)
        childNode = childNode.GetNext


# 创建表格并填写表头
wb = openpyxl.Workbook()  # 创建一个新的工作簿
ws = wb.active  # 获取当前的工作表
ws.title = 表格名称  # 给工作表命名
ws['A1'] = '零件名称'  # 在A1单元格写入零件名称
ws['B1'] = 'comX'  # 在B1~D1单元格写入质心X、Y、Z
ws['C1'] = 'comY'
ws['D1'] = 'comZ'
ws['E1'] = '体积'
ws['F1'] = '表面积'
ws['G1'] = '质量'
ws['H1'] = 'Ixx'  # 惯性矩Ixx，单位是千克·平方米
ws['I1'] = 'Ixy'
ws['J1'] = 'Ixz'
ws['K1'] = 'Iyx'
ws['L1'] = 'Iyy'
ws['M1'] = 'Iyz'
ws['N1'] = 'Izx'
ws['O1'] = 'Izy'
ws['P1'] = 'Izz'



# 连接Solidworks
# 建立com连接,如只有一个版本,可以只写"SldWorks.Application"
swApp = win32com.client.Dispatch(f'SldWorks.Application.{版本号 - 1992}')
swApp.CommandInProgress = True  # 提升API交互效率
swApp.Visible = True  # 显示SolidWorks界面
Part = swApp.ActiveDoc  # 获取当前打开的装配体

# 遍历整个装配体，讲所有零件的名称储存在rootNode中（格式无法直接使用）
featureMgr = Part.FeatureManager
rootNode = featureMgr.GetFeatureTreeRootItem2(constants.swFeatMgrPaneBottom)
traverse_node(rootNode)

# 计算出rootNode中所包含的零件数目
swCompColl = Part.GetComponents(False)  # 获取装配体的组件集合

# 逐个获取其中所包含的零件以及整个装配体
vModels = swApp.GetDocuments



# 将所有零件的名字填写到第一列
for i in range(len(selectedpart)):  # 遍历rootNode中的元素个数
    零件的名称 = selectedpart[i]
    print(i)
    print(零件的名称)
    ws.cell(row=i + 2, column=1).value = 零件的名称
ws.cell(row=len(selectedpart) + 2, column=1).value = "together"


i = 0
print(f"已打开文件数量为：{len(vModels)}")
for art in vModels:
    # print(f"文件路径名：{art.GetPathName}")
    # print(art.GetMassProperties)
    零件的名称 = art.GetPathName
    print(零件的名称)

    mass_property = art.Extension.CreateMassProperty
    #    arg1 = win32com.client.VARIANT(pythoncom.VT_BSTR, None)
    #    coord_sys = Part.Extension.GetCoordinateSystemTransformByName(arg1)
    arg1 = win32com.client.VARIANT(pythoncom.VT_DISPATCH, None)
    mass_property.SetCoordinateSystem(arg1)

    # ws.cell(row=i + 2, column=1).value = 零件的名称

    com = mass_property.CenterOfMass
    comX = com[0]
    ws.cell(row=i + 2, column=2).value = com[0]
    comY = com[1]
    ws.cell(row=i + 2, column=3).value = com[1]
    comZ = com[2]
    ws.cell(row=i + 2, column=4).value = com[2]

    V = mass_property.Volume
    ws.cell(row=i + 2, column=5).value = mass_property.Volume
    surface = mass_property.SurfaceArea
    ws.cell(row=i + 2, column=6).value = mass_property.SurfaceArea
    m = mass_property.Mass
    print(m)
    ws.cell(row=i + 2, column=7).value = mass_property.Mass
#0是对坐标系，1是对质心

    I = mass_property.GetMomentOfInertia(0)

    Ixx = I[0]
    Ixy = I[1]
    Ixz = I[2]
    Iyx = I[3]
    Iyy = I[4]
    Iyz = I[5]
    Izx = I[6]
    Izy = I[7]
    Izz = I[8]

    # print("here")
    ws.cell(row=i + 2, column=8).value = I[0]
    ws.cell(row=i + 2, column=9).value = I[1]
    ws.cell(row=i + 2, column=10).value = I[2]
    ws.cell(row=i + 2, column=11).value = I[3]
    ws.cell(row=i + 2, column=12).value = I[4]
    ws.cell(row=i + 2, column=13).value = I[5]
    ws.cell(row=i + 2, column=14).value = I[6]
    ws.cell(row=i + 2, column=15).value = I[7]
    ws.cell(row=i + 2, column=16).value = I[8]


    i = i + 1
wb.save(表格名称 + ".xlsx")  # 保存工作簿

wb.close()  # 关闭工作簿
